"""Spreadsheet loaders (Excel via openpyxl, CSV via stdlib).

Excel: one ParsedPart per worksheet. Each sheet is serialized as tab-separated
rows — the first non-empty row is treated as the header and repeated as context
isn't necessary because the header row already leads the text. Sheet name is
emitted as a level-1 heading so the chunker keeps sheets as section boundaries
("Workbook > Q3 Revenue").

CSV: a single ParsedPart (one logical sheet). Same tab-serialization so the two
spreadsheet kinds land in identical shape downstream.

Both flatten cells to `str`; numbers/dates come back via their str() form. This
is deliberately lossy — we're indexing a spreadsheet for semantic search and
grounded chat, not reconstructing formulas. A dedicated table-aware parser
(Docling) is the heavyweight upgrade path if cell-level citation ever matters.
"""

import csv
from collections.abc import Iterable, Iterator
from pathlib import Path

from pynote_core.parsers.types import ParsedPart

# Cap serialized rows per sheet. A 200k-row export would otherwise produce one
# multi-megabyte part that blows the chunker and embedding budget; the first N
# rows carry the schema + a representative sample, which is what retrieval needs.
_MAX_ROWS = 2_000
# Skip rows that are entirely empty after stripping — spacer rows are common in
# hand-made spreadsheets and add nothing but noise to the embedding.
_TRUNCATION_NOTE = "[… additional rows truncated …]"


def _cell(value: object) -> str:
    """Render one cell to text. None → '' so empty cells collapse cleanly."""
    if value is None:
        return ""
    return str(value).strip()


def _serialize_rows(rows: Iterable[tuple[object, ...]]) -> str:
    """Tab-join each row, drop all-empty rows, and cap at _MAX_ROWS.

    Returns the sheet's text body. A truncation marker is appended when the row
    cap is hit so downstream readers know the view is partial.
    """
    lines: list[str] = []
    truncated = False
    for i, row in enumerate(rows):
        if i >= _MAX_ROWS:
            truncated = True
            break
        cells = [_cell(v) for v in row]
        if not any(cells):
            continue
        lines.append("\t".join(cells))
    if truncated:
        lines.append(_TRUNCATION_NOTE)
    return "\n".join(lines)


def parse_spreadsheet(path: Path) -> Iterator[ParsedPart]:
    """Yield one ParsedPart per worksheet in an .xlsx workbook.

    `read_only=True` streams rows without loading the whole sheet into memory;
    `data_only=True` returns the last-cached formula *values* rather than the
    formula strings, which is what a reader actually wants to search over.
    """
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        for ordinal, ws in enumerate(wb.worksheets):
            body = _serialize_rows(ws.iter_rows(values_only=True))
            # Sheet name leads the text so the heading offset is a genuine 0.
            text = f"{ws.title}\n{body}" if body else ws.title
            yield ParsedPart(
                ordinal=ordinal,
                text=text,
                headings=[{"text": ws.title, "level": 1, "start": 0}],
            )
    finally:
        wb.close()


def parse_csv(path: Path) -> Iterator[ParsedPart]:
    """Yield a single ParsedPart for a CSV file.

    Dialect is sniffed from a sample so tab- and semicolon-delimited exports
    parse correctly; on failure we fall back to the comma default.
    """
    with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
        sample = f.read(8192)
        f.seek(0)
        try:
            dialect: type[csv.Dialect] | csv.Dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel  # comma-delimited default
        reader = csv.reader(f, dialect)
        body = _serialize_rows(reader)

    yield ParsedPart(
        ordinal=0,
        text=body,
        headings=None,  # a bare CSV has no sheet name / title to anchor
    )
